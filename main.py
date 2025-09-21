#from __future__ import print_function
import os
import re
import pandas as pd
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
import email
import pickle
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

# Si modifica los alcances, elimina el archivo token.pickle
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

# Config / util

def cargar_variables():
    load_dotenv()
    # Soportar EMAIL_PASS o EMAIL_PASSWORD por compatibilidad
    user = os.getenv("EMAIL_USER")
    password = os.getenv("EMAIL_PASS") or os.getenv("EMAIL_PASSWORD")
    smtp_server = os.getenv("EMAIL_SMTP")
    smtp_port = int(os.getenv("EMAIL_PORT", 587))
    if not all([user, password, smtp_server]):
        raise EnvironmentError("Faltan variables de entorno. Define EMAIL_USER, EMAIL_PASS (o EMAIL_PASSWORD) y EMAIL_SMTP.")
    return user, password, smtp_server, smtp_port

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def validar_email(email):
    if not isinstance(email, str) or not email.strip():
        return False
    return bool(EMAIL_REGEX.match(email.strip()))


# Lectura / plantilla

def leer_csv(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe el archivo CSV: {path}")
    try:
        # Intentar con UTF-8 primero
        return pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        # Si falla (ej: archivo guardado en Windows/Excel con acentos)
        return pd.read_csv(path, encoding="latin1")

def leer_plantilla(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe la plantilla: {path}")
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


# Envío de email

def enviar_correo(user, password, smtp_server, smtp_port, destinatario, asunto, cuerpo):

    try:
        msg = EmailMessage()
        msg['From'] = user
        msg['To'] = destinatario
        msg['Subject'] = asunto
        msg.set_content(cuerpo)

        with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(user, password)
            server.send_message(msg)

        return 'SENT', ''
    except Exception as e:
        return 'ERROR', str(e)


# Logging (archivo envios_email.log con separador '|') para mayor organizacion

def registrar_log(log_file, fecha, destinatario, estado, mensaje_error):
    with open(log_file, 'a', encoding='utf-8') as f:
        # Reemplazar saltos de línea en mensaje_error para mantener formato por filas
        msg_clean = (mensaje_error or '').replace('\n', ' ').replace('\r', '')
        f.write(f"{fecha}|{destinatario}|{estado}|{msg_clean}\n")

def buscar_rebotes(service):
    query = 'subject:(Delivery OR "Mail Delivery" OR Undelivered OR "Returned mail" OR failure OR "Delivery Status")'
    results = service.users().messages().list(userId='me', q=query, maxResults=50).execute()
    messages = results.get('messages', [])

    rebotes = []

    for msg in messages:
        msg_data = service.users().messages().get(userId='me', id=msg['id'], format='raw').execute()
        raw = base64.urlsafe_b64decode(msg_data['raw'].encode('ASCII'))
        email_message = email.message_from_bytes(raw)

        subject = email_message['Subject']
        from_ = email_message['From']

        cuerpo = ""
        for part in email_message.walk():
            if part.get_content_type() == 'text/plain':
                cuerpo = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                break

        re_email = re.search(r'[\w\.-]+@[\w\.-]+', cuerpo)
        email_rebotado = re_email.group(0) if re_email else "Desconocido"

        rebotes.append({'email': email_rebotado, 'subject': subject, 'from': from_})

    return rebotes

def limpiar_log(log_file):
    if os.path.exists(log_file):
        os.remove(log_file)


# Reporte Excel (reporte_envios.xlsx)

def reporte_excel(log_file, reporte_file):
    if not os.path.exists(log_file):
        raise FileNotFoundError(f"No existe el log: {log_file}")

    df = pd.read_csv(log_file, sep='|', names=['Fecha', 'Destinatario', 'Estado', 'Mensaje_error'])

    # NORMALIZAMOS POR DESTINATARIO: El peor estado tiene prioridad (ERROR > SENT)
    df.sort_values(by='Fecha', ascending=True, inplace=True)

    def estado_final(grupo):
        if 'ERROR' in grupo['Estado'].values:
            return grupo[grupo['Estado'] == 'ERROR'].iloc[-1]
        return grupo.iloc[-1]

    df_final = df.groupby('Destinatario', group_keys=False).apply(estado_final).reset_index(drop=True)

    total = len(df_final)
    exitos = int((df_final['Estado'] == 'SENT').sum())
    errores = int((df_final['Estado'] == 'ERROR').sum())
    porc_exitos = round(exitos / total * 100, 2) if total else 0
    porc_errores = round(errores / total * 100, 2) if total else 0

    # Crear Excel reporte_envios.xlsx
    wb = Workbook()
    ws_listado = wb.active
    ws_listado.title = "Listado"

    for fila in dataframe_to_rows(df_final, index=False, header=True):
        ws_listado.append(fila)

    #  Gráficas - Solo PieChart -> Grafica de pastel
    ws_graficas = wb.create_sheet("Graficas")
    ws_graficas.append(["Estado", "Cantidad"])
    ws_graficas.append(["SENT", exitos])
    ws_graficas.append(["ERROR", errores])

    data_pie = Reference(ws_graficas, min_col=2, min_row=1, max_row=3)
    labels_pie = Reference(ws_graficas, min_col=1, min_row=2, max_row=3)

    pie = PieChart()
    pie.title = "Distribución de Estados"
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(labels_pie)
    ws_graficas.add_chart(pie, "D2")

    # Resumen Final
    ws_resumen = wb.create_sheet("Resumen")
    resumen = [
        ["Métrica", "Valor"],
        ["Total de destinatarios únicos", total],
        ["Envíos exitosos (SENT)", exitos],
        ["Errores detectados (ERROR)", errores],
        ["Porcentaje de éxito (%)", porc_exitos],
        ["Porcentaje de error (%)", porc_errores],
    ]
    for fila in resumen:
        ws_resumen.append(fila)

    wb.save(reporte_file)
    print(f"📊 Reporte generado: {reporte_file}")

def procesar_rebotes_si_gmail(user, service, log_file):
    if '@gmail.com' in user.lower():
        print("Buscando rebotes en Gmail...")
        rebotes = buscar_rebotes(service)
        for rebote in rebotes:
            email_rebotado = rebote['email']
            fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            registrar_log(log_file, fecha, email_rebotado, 'ERROR', 'Rebote detectado en Gmail')
            print(f"[{fecha}] {email_rebotado} -> ERROR (rebote detectado)")
    else:
        print("Rebotes no verificados: la cuenta no es de Gmail.")

def main():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('gmail', 'v1', credentials=creds)

    print("Iniciando proceso de envíos...")
    user, password, smtp_server, smtp_port = cargar_variables()

    clientes_csv = 'clientes.csv'
    plantilla_file = 'plantilla.txt'
    log_file = 'envios_email.log'
    reporte_file = 'reporte_envios.xlsx'

    limpiar_log(log_file)

    clientes_df = leer_csv(clientes_csv)
    plantilla_text = leer_plantilla(plantilla_file)

    if 'empresa' not in clientes_df.columns or 'email' not in clientes_df.columns:
        raise ValueError("El CSV debe contener las columnas 'empresa' y 'email'.")

    for _, row in clientes_df.iterrows():
        empresa = str(row.get('empresa', '')).strip()
        email_ = str(row.get('email', '')).strip()
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if not validar_email(email_):
            registrar_log(log_file, fecha, email_ or 'NO_EMAIL', 'ERROR', 'Formato de email inválido')
            print(f"[{fecha}] SKIP/ERROR -> {email_} (formato inválido)")
            continue

        cuerpo = plantilla_text.replace('{empresa}', empresa)
        estado, error = enviar_correo(user, password, smtp_server, smtp_port, email_, f"Asunto para {empresa}", cuerpo)
        registrar_log(log_file, fecha, email_, estado, error)
        print(f"[{fecha}] {email_} -> {estado}")

    #  Verifica rebotes si la cuenta es de Gmail
    procesar_rebotes_si_gmail(user, service, log_file)

    #  Genera el reporte final
    reporte_excel(log_file, reporte_file)
    print("Proceso finalizado. Revisa:", log_file, reporte_file)

if __name__ == "__main__":
    main()